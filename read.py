import pandas as pd
from pandas import option_context
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_excel('sales.xlsx')[['Gender', 'Product line', 'Total']]
#print(excel_file)
report_table = excel_file.pivot_table(index='Gender',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round(0)
print(report_table)

report_table.to_excel('report_2021.xlsx',
                      sheet_name='Report',
                      startrow=4)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# barchart
barchart = BarChart()
#locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row)
# adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.height = 15 # default is 7.5
barchart.width = 20 # default is 15
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 2

worksheet = wb.active
for col in worksheet.columns:
     max_length = 0
     column = col[0].column_letter # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     worksheet.column_dimensions[column].width = adjusted_width

wb.save('report_2021.xlsx')